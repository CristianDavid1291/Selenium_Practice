import openpyxl

class Excel_Functions():

    def __init__(self,driver):
        self.driver = driver

    def getCountRow(file,path,sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file,sheetName):
         workbook = openpyxl.load_workbook(file)
         sheet = workbook[sheetName]
         return (sheet.max_column)

    def readData(file,path,sheetName,numRow,numColum):
         workbook = openpyxl.load_workbook(path)
         sheet = workbook[sheetName]
         return sheet.cell(row=numRow, column=numColum).value
    
    def writeData(file,path,sheetName,numRow,numColumn,data):
         workbook = openpyxl.load_workbook(path)
         sheet = workbook[sheetName]
         sheet.cell(row=numRow,column=numColumn).value = data
         workbook.save(path)